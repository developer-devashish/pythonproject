import tkinter as tk
import time
import random
from openpyxl import Workbook, load_workbook
import os

root = tk.Tk()
root.title("BFF (Brooklyn Food Factory)")
root.geometry("900x600")


# ---------------- SAVE TO EXCEL ----------------
def save_to_excel():
    file_name = "Cafe_Bill_Record.xlsx"

    # If excel exists → load it, else create it
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Bill No", "Date & Time", "Coffee Qty", "Tea Qty", "Cake Qty",
                      "Pastry Qty", "Sandwich Qty", "Total", "Tax", "Grand Total"])

    sheet.append([
        bill_no_var.get(),
        date_var.get(),
        entry_coffee.get() or 0,
        entry_tea.get() or 0,
        entry_cake.get() or 0,
        entry_pastry.get() or 0,
        entry_sandwich.get() or 0,
        total_var.get(),
        tax_var.get(),
        grand_total_var.get()
    ])

    wb.save(file_name)


# ---------------- RESET ----------------
def reset():
    entry_coffee.set("")
    entry_tea.set("")
    entry_pastry.set("")
    entry_sandwich.set("")
    entry_cake.set("")
    total_var.set("")
    tax_var.set("")
    grand_total_var.set("")
    bill_no_var.set("")
    date_var.set("")


# ---------------- RECEIPT WINDOW ----------------
def generate_receipt():
    receipt_window = tk.Toplevel(root)
    receipt_window.title("Receipt")
    receipt_window.geometry("450x550")

    receipt_text = tk.Text(receipt_window, font=("Courier", 12))
    receipt_text.pack(fill=tk.BOTH, expand=True)

    # Get values
    coffee = int(entry_coffee.get() or 0)
    tea = int(entry_tea.get() or 0)
    cake = int(entry_cake.get() or 0)
    pastry = int(entry_pastry.get() or 0)
    sandwich = int(entry_sandwich.get() or 0)

    bill_no = bill_no_var.get()
    date_time = date_var.get()

    receipt_text.insert(tk.END, "   BFF (Brooklyn Food Factory)\n")
    receipt_text.insert(tk.END, "-------------------------------------\n")
    receipt_text.insert(tk.END, f" Bill No: {bill_no}\n")
    receipt_text.insert(tk.END, f" Date & Time: {date_time}\n")
    receipt_text.insert(tk.END, "-------------------------------------\n")
    receipt_text.insert(tk.END, " Item           Qty    Price\n")
    receipt_text.insert(tk.END, "-------------------------------------\n")

    if coffee: receipt_text.insert(tk.END, f" Coffee        {coffee}     {coffee * 50}\n")
    if tea: receipt_text.insert(tk.END, f" Tea           {tea}     {tea * 30}\n")
    if cake: receipt_text.insert(tk.END, f" Cake          {cake}     {cake * 80}\n")
    if pastry: receipt_text.insert(tk.END, f" Pastry        {pastry}     {pastry * 60}\n")
    if sandwich: receipt_text.insert(tk.END, f" Sandwich      {sandwich}     {sandwich * 100}\n")

    receipt_text.insert(tk.END, "-------------------------------------\n")
    receipt_text.insert(tk.END, f" Total:               {total_var.get()}\n")
    receipt_text.insert(tk.END, f" Tax (5%):            {tax_var.get()}\n")
    receipt_text.insert(tk.END, f" Grand Total:         {grand_total_var.get()}\n")
    receipt_text.insert(tk.END, "-------------------------------------\n")
    receipt_text.insert(tk.END, "   Thank you! Visit Again 😊\n")


# ---------------- TOTAL CALCULATION ----------------
def total():
    coffee = int(entry_coffee.get() or 0)
    tea = int(entry_tea.get() or 0)
    cake = int(entry_cake.get() or 0)
    pastry = int(entry_pastry.get() or 0)
    sandwich = int(entry_sandwich.get() or 0)

    total_cost = (coffee * 50) + (tea * 30) + (cake * 80) + (pastry * 60) + (sandwich * 100)
    tax = total_cost * 0.05
    grand_total = total_cost + tax

    total_var.set(f"{total_cost:.2f}")
    tax_var.set(f"{tax:.2f}")
    grand_total_var.set(f"₹{grand_total:.2f}")

    bill_no_var.set(random.randint(1000, 9999))
    date_var.set(time.strftime("%d-%m-%Y  %I:%M %p"))

    generate_receipt()
    save_to_excel()  # <-- Auto Save to Excel


# ------------------------------------------------------------
# --------------------------- GUI -----------------------------
# ------------------------------------------------------------

title_label = tk.Label(root, text="☕ BFF (Brooklyn Food Factory) ☕",
                       font=('Times New Roman', 28, 'bold'),
                       bg='brown', fg='white', pady=10)
title_label.pack(fill=tk.X)

# ---------- MENU FRAME ----------
menu_frame = tk.Frame(root, bd=10, relief=tk.GROOVE)
menu_frame.place(x=20, y=100, width=400, height=380)

tk.Label(menu_frame, text="Menu", font=('Arial', 20, 'bold')).grid(row=0, column=0, columnspan=2, pady=10)

tk.Label(menu_frame, text="Coffee (₹50)").grid(row=1, column=0, sticky="w", padx=10)
tk.Label(menu_frame, text="Tea (₹30)").grid(row=2, column=0, sticky="w", padx=10)
tk.Label(menu_frame, text="Cake (₹80)").grid(row=3, column=0, sticky="w", padx=10)
tk.Label(menu_frame, text="Pastry (₹60)").grid(row=4, column=0, sticky="w", padx=10)
tk.Label(menu_frame, text="Sandwich (₹100)").grid(row=5, column=0, sticky="w", padx=10)

entry_coffee = tk.StringVar()
entry_tea = tk.StringVar()
entry_cake = tk.StringVar()
entry_pastry = tk.StringVar()
entry_sandwich = tk.StringVar()

tk.Entry(menu_frame, textvariable=entry_coffee, width=10).grid(row=1, column=1)
tk.Entry(menu_frame, textvariable=entry_tea, width=10).grid(row=2, column=1)
tk.Entry(menu_frame, textvariable=entry_cake, width=10).grid(row=3, column=1)
tk.Entry(menu_frame, textvariable=entry_pastry, width=10).grid(row=4, column=1)
tk.Entry(menu_frame, textvariable=entry_sandwich, width=10).grid(row=5, column=1)


# -------- BILL FRAME --------
bill_frame = tk.Frame(root, bd=10, relief=tk.GROOVE)
bill_frame.place(x=450, y=100, width=400, height=380)

tk.Label(bill_frame, text="Bill Details", font=("Arial", 20, "bold")).grid(row=0, column=0, columnspan=2, pady=10)

tk.Label(bill_frame, text="Total: ").grid(row=1, column=0, sticky="w", padx=10)
tk.Label(bill_frame, text="Tax (5%): ").grid(row=2, column=0, sticky="w", padx=10)
tk.Label(bill_frame, text="Grand Total: ").grid(row=3, column=0, sticky="w", padx=10)
tk.Label(bill_frame, text="Bill No: ").grid(row=4, column=0, sticky="w", padx=10)
tk.Label(bill_frame, text="Date & Time: ").grid(row=5, column=0, sticky="w", padx=10)

total_var = tk.StringVar()
tax_var = tk.StringVar()
grand_total_var = tk.StringVar()
bill_no_var = tk.StringVar()
date_var = tk.StringVar()

tk.Label(bill_frame, textvariable=total_var).grid(row=1, column=1, sticky="e")
tk.Label(bill_frame, textvariable=tax_var).grid(row=2, column=1, sticky="e")
tk.Label(bill_frame, textvariable=grand_total_var).grid(row=3, column=1, sticky="e")
tk.Label(bill_frame, textvariable=bill_no_var).grid(row=4, column=1, sticky="e")
tk.Label(bill_frame, textvariable=date_var).grid(row=5, column=1, sticky="e")


# ---------- BUTTONS ----------
button_frame = tk.Frame(root, bd=10, relief=tk.GROOVE)
button_frame.place(x=20, y=500, width=830, height=80)

tk.Button(button_frame, text="Generate Bill", bg="green", fg="white",
          font=('Arial', 14, 'bold'), command=total).grid(row=0, column=0, padx=50)

tk.Button(button_frame, text="Reset", bg="orange", fg="white",
          font=('Arial', 14, 'bold'), command=reset).grid(row=0, column=1, padx=50)

tk.Button(button_frame, text="Exit", bg="red", fg='white',
          font=('Arial', 14, 'bold'), command=root.destroy).grid(row=0, column=2, padx=50)

root.mainloop()
